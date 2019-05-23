import openpyxl
from html_table_extractor.extractor import Extractor
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import re

class Table2xlsx(object):
    
    @classmethod
    def html_table_converter(self, offer_table):

        extractor = Extractor(offer_table)
        extractor.parse()
        table_list = extractor.return_list()
        return table_list
    
    @classmethod
    def table_to_xlsx(self, workbook, sheetname, table_list):
        wb = openpyxl.load_workbook(workbook)
        sheet = wb[sheetname]
        if sheet.max_row>1:
            next_row = sheet.max_row+2
        else:
            next_row = 1
        fontObj1 = Font(bold=True)
        for i in range(len(table_list)):
            for j in range(len(table_list[i])):
                sheet.cell(row=next_row+i, column=j+1).value = table_list[i][j]
                if i==0:
                    sheet.cell(row=next_row+i, column=j+1).font = fontObj1
        
        for col in sheet.columns:
             max_length = 0
             column = col[0].column
             
             for cell in col:
                 try: 
                     if len(str(cell.value)) > max_length:
                         max_length = len(cell.value)
                 except:
                     pass
             adjusted_width = max_length + 2
             sheet.column_dimensions[get_column_letter(column)].width = adjusted_width       
        wb.save(workbook)

    @classmethod
    def fill_missing(self, category_detail, list_text, link, table_list):
        #regex area
        catreg = re.compile(r"\w*\sHotels|\w*\sFlights|\w*\sCabs")
        valreg = re.compile(r"valid.*\d+")
        conreg = re.compile(r"\s\d\sbooking.+|once.+")
        #regex area
        
        standard = ["Platform","Coupon Code","Category","Offer Details",
                    "Minimum Booking Amount (INR)","Booking Channel",
                    "Applicable Banks","Validity","Constraints","Offer Link"]
        for head in range(len(table_list[0])):
            index =[]
            if table_list[0][head] not in standard:
                index.append(head)
        for i in index:
            table_list[0].pop(i)
            for li in range(1,len(table_list)):
                table_list[li].pop(i)
        if "Category" not in table_list[0]:
            table_list[0].insert(1,"Category")            
            category_head = catreg.search(category_detail)
            if category_head:
                category_head = category_head.group(0)
            else:
                category_head = "Uncategorized"
            for index in range(1,len(table_list)):
                table_list[index].insert(1,category_head)

        table_list[0].insert(0,"Platform")
        for index in range(1,len(table_list)):
            table_list[index].insert(0,"MakeMyTrip")
            
        if "Validity" not in table_list[0]:
            table_list[0].append("Validity") 
            validity = valreg.search(list_text)
            if validity:
                validity = validity.group(0)
            else:
                validity = "N/A"
            for index in range(1,len(table_list)):
                table_list[index].append(validity)

        table_list[0].append("Constraints")
        constraints = conreg.findall(list_text)
        constraint = ""
        if constraints:
            for con in constraints:
                constraint += con + "," 
        else:
            constraint = "N/A"
        for index in range(1,len(table_list)):
            table_list[index].append(constraint)
            
        table_list[0].append("Offer Link")
        link = '=HYPERLINK("{}", "{}")'.format(link, "Link")
        for index in range(1,len(table_list)):
            table_list[index].append(link)

        return table_list
